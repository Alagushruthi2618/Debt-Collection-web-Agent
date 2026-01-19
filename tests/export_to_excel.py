"""
Script to export test results from JSON to Excel format.
"""
import json
import sys
from pathlib import Path

try:
    import pandas as pd
except ImportError:
    print("pandas is required. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd


def export_test_results_to_excel(json_file_path: str, excel_file_path: str = None):
    """
    Export test results from JSON to Excel format.
    
    Args:
        json_file_path: Path to the test_results.json file
        excel_file_path: Path for the output Excel file (optional)
    """
    # Read JSON file
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Prepare data for Excel
    rows = []
    
    for category, category_data in data.items():
        total = category_data.get('total', 0)
        passed = category_data.get('passed', 0)
        failed = category_data.get('failed', 0)
        pass_rate = category_data.get('pass_rate', 0.0)
        
        # Add summary row for each category
        rows.append({
            'Category': category.upper(),
            'Test Case Input': f'SUMMARY - Total: {total}, Passed: {passed}, Failed: {failed}, Pass Rate: {pass_rate:.2%}',
            'Status': 'PASS' if failed == 0 else 'FAIL',
            'Detected Status': '',
            'Error': '',
            'Total Tests': total,
            'Passed': passed,
            'Failed': failed,
            'Pass Rate': f'{pass_rate:.2%}'
        })
        
        # Add individual test cases
        test_cases = category_data.get('test_cases', [])
        for idx, test_case in enumerate(test_cases, 1):
            rows.append({
                'Category': category,
                'Test Case Input': test_case.get('input', ''),
                'Status': 'PASS' if test_case.get('passed', False) else 'FAIL',
                'Detected Status': test_case.get('detected_status', ''),
                'Error': test_case.get('error', ''),
                'Total Tests': '',
                'Passed': '',
                'Failed': '',
                'Pass Rate': ''
            })
        
        # Add empty row between categories for better readability
        rows.append({
            'Category': '',
            'Test Case Input': '',
            'Status': '',
            'Detected Status': '',
            'Error': '',
            'Total Tests': '',
            'Passed': '',
            'Failed': '',
            'Pass Rate': ''
        })
    
    # Create DataFrame
    df = pd.DataFrame(rows)
    
    # Determine output file path
    if excel_file_path is None:
        json_path = Path(json_file_path)
        excel_file_path = json_path.parent / f"{json_path.stem}.xlsx"
    
    # Write to Excel with formatting
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Test Results', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Test Results']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)  # Cap at 100 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Style the header row
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Style summary rows (rows with 'SUMMARY' in Test Case Input)
        summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        bold_font = Font(bold=True)
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            test_input = row[1].value if row[1].value else ""
            if 'SUMMARY' in str(test_input):
                for cell in row:
                    cell.fill = summary_fill
                    cell.font = bold_font
        
        # Style pass/fail cells
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for row in worksheet.iter_rows(min_row=2):
            status_cell = row[2]  # Status column
            if status_cell.value == 'PASS':
                status_cell.fill = pass_fill
            elif status_cell.value == 'FAIL':
                status_cell.fill = fail_fill
    
    print(f"Test results exported to: {excel_file_path}")
    print(f"  Total categories: {len(data)}")
    total_tests = sum(cat_data.get('total', 0) for cat_data in data.values())
    print(f"  Total test cases: {total_tests}")


if __name__ == "__main__":
    # Default paths
    script_dir = Path(__file__).parent
    json_file = script_dir / "test_results.json"
    
    # Allow command line argument for JSON file path
    if len(sys.argv) > 1:
        json_file = Path(sys.argv[1])
    
    # Allow command line argument for Excel output path
    excel_file = None
    if len(sys.argv) > 2:
        excel_file = sys.argv[2]
    
    if not json_file.exists():
        print(f"Error: JSON file not found at {json_file}")
        sys.exit(1)
    
    export_test_results_to_excel(str(json_file), excel_file)
